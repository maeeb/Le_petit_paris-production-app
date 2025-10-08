from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone 
from django.db.models import Sum, Count, Q, Avg
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.core.paginator import Paginator
from datetime import datetime, timedelta
from .models import ProductionRecord, Alerte, LigneProduction, Produit, Equipe, UserProfile
from .forms import ProductionForm
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.contrib.auth.models import User
import json
from django.db.models import Max



# =====================================================
# FONCTIONS UTILITAIRES ET CALCULS KPI
# =====================================================
def operateur_login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('operator_dashboard')  # redirige vers le dashboard opérateur
        else:
            return render(request, 'operateur_login.html', {'error': 'Identifiants incorrects'})
    return render(request, 'operateur_login.html')
def calculate_kpis(records):
    """Calcule tous les KPIs à partir des enregistrements fournis"""
    if not records.exists():
        return {
            'taux_productivite': 0,
            'taux_defauts': 0,
            'temps_arret_moyen': 0,
            'palettes_par_heure': 0,
            'efficacite_globale': 0,
            'total_palettes': 0,
            'total_dechets': 0,
            'total_arrets': 0,
            'total_palettes_non_conformes': 0,
            'total_palettes_conformes': 0,
            'taux_conformite': 0,
            'nombre_records': 0,
            'productivite_moyenne': 0
        }
    
    aggregations = records.aggregate(
        total_palettes=Sum('palettes_produites'),
        total_palettes_non_conformes=Sum('palettes_non_conformes'),
        total_dechets=Sum('dechets_boites'),
        total_arrets=Sum('duree_arret_minutes'),
        avg_arrets=Avg('duree_arret_minutes'),
        count_records=Count('id')
    )
    
    total_palettes = aggregations['total_palettes'] or 0
    total_palettes_nc = aggregations['total_palettes_non_conformes'] or 0
    total_dechets = aggregations['total_dechets'] or 0
    total_arrets = aggregations['total_arrets'] or 0
    avg_arrets = aggregations['avg_arrets'] or 0
    count_records = aggregations['count_records'] or 0
    
    total_palettes_conformes = total_palettes - total_palettes_nc
    
    taux_conformite = (total_palettes_conformes / total_palettes * 100) if total_palettes > 0 else 0
    palettes_par_heure = total_palettes / count_records if count_records > 0 else 0
    objectif_theorique = count_records * 8
    taux_productivite = (total_palettes / objectif_theorique * 100) if objectif_theorique > 0 else 0
    production_totale_boites = total_palettes * 25
    taux_defauts = (total_dechets / production_totale_boites * 100) if production_totale_boites > 0 else 0
    efficacite_globale = (taux_productivite * taux_conformite / 100) if taux_productivite > 0 and taux_conformite > 0 else 0
    productivite_moyenne = total_palettes / count_records if count_records > 0 else 0
    
    return {
        'taux_productivite': round(taux_productivite, 2),
        'taux_defauts': round(taux_defauts, 2),
        'temps_arret_moyen': round(avg_arrets, 2),
        'palettes_par_heure': round(palettes_par_heure, 2),
        'efficacite_globale': round(efficacite_globale, 2),
        'total_palettes': total_palettes,
        'total_dechets': total_dechets,
        'total_arrets': total_arrets,
        'total_palettes_non_conformes': total_palettes_nc,
        'total_palettes_conformes': total_palettes_conformes,
        'taux_conformite': round(taux_conformite, 2),
        'nombre_records': count_records,
        'productivite_moyenne': round(productivite_moyenne, 2)
    }


def calculate_team_productivity():
    """Calcule la productivité par équipe pour aujourd'hui"""
    today = timezone.now().date()
    
    records = ProductionRecord.objects.filter(
        date_heure__date=today
    ).select_related('equipe', 'operateur')
    
    team_stats = {}
    for record in records:
        team_key = record.equipe.nom if record.equipe else f"Équipe {record.operateur.username}"
        if team_key not in team_stats:
            team_stats[team_key] = {
                'equipe_nom': team_key,
                'horaire': record.equipe.horaire if record.equipe else 'Non défini',
                'records': [],
                'operateurs': set()
            }
        team_stats[team_key]['records'].append(record)
        team_stats[team_key]['operateurs'].add(record.operateur.username)
    
    team_productivity = []
    for team_name, team_data in team_stats.items():
        team_records = ProductionRecord.objects.filter(
            id__in=[r.id for r in team_data['records']]
        )
        kpis = calculate_kpis(team_records)
        
        team_info = {
            'equipe_nom': team_name,
            'horaire': team_data['horaire'],
            'nombre_operateurs': len(team_data['operateurs']),
            'operateurs': list(team_data['operateurs']),
            'total_palettes': kpis['total_palettes'],
            'total_conformes': kpis['total_palettes_conformes'],
            'total_non_conformes': kpis['total_palettes_non_conformes'],
            'palettes_par_heure': kpis['palettes_par_heure'],
            'taux_conformite': kpis['taux_conformite'],
            'temps_travaille': 480,  # 8 heures par défaut
        }
        team_productivity.append(team_info)
    
    team_productivity.sort(key=lambda x: x['total_palettes'], reverse=True)
    return team_productivity


def determine_current_shift():
    """Détermine l'horaire d'équipe actuel"""
    current_hour = timezone.localtime(timezone.now()).hour
    
    if 6 <= current_hour < 14:
        return '6-14', "Équipe du Matin"
    elif 14 <= current_hour < 22:
        return '14-22', "Équipe de l'Après-midi"
    else:
        return '22-6', "Équipe de Nuit"


# =====================================================
# VUES D'AUTHENTIFICATION
# =====================================================

@csrf_protect
def login_view(request):
    """Vue de connexion avec gestion des horaires d'équipe"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        if username and password:
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                
                # Vérifier l'accès après connexion
                if user.is_superuser or user.is_staff:
                    return redirect('admin_dashboard')
                elif hasattr(user, 'userprofile') and user.userprofile.can_access_now():
                    return redirect('operator_dashboard')
                else:
                    # Rediriger vers page de blocage
                    return redirect('shift_blocked')
            else:
                messages.error(request, 'Identifiants incorrects')
        else:
            messages.error(request, 'Veuillez remplir tous les champs')
    
    return render(request, 'production/login.html')


def logout_view(request):
    """Déconnexion avec message de confirmation"""
    logout(request)
    messages.success(request, 'Vous avez été déconnecté avec succès')
    return redirect('login')


def shift_blocked(request):
    """Page affichée quand l'accès est bloqué selon l'horaire"""
    current_hour = datetime.now().hour
    current_shift, shift_name = determine_current_shift()
    
    user_shift = None
    user_next_access = None
    
    if request.user.is_authenticated and hasattr(request.user, 'userprofile'):
        profile = request.user.userprofile
        # CORRECTION : récupérer la valeur brute
        user_shift = profile.horaire_equipe
        
        # CORRECTION : Déterminer le prochain accès selon TOUS les formats
        if profile.horaire_equipe in ['6-14', '06-14']:
            user_next_access = "06h00 à 14h00"
        elif profile.horaire_equipe in ['14-22']:
            user_next_access = "14h00 à 22h00"  
        elif profile.horaire_equipe in ['22-6', '22-06']:
            user_next_access = "22h00 à 06h00"
        else:
            user_next_access = "Horaire non défini"
    
    context = {
        'current_hour': current_hour,
        'current_shift': shift_name,
        'user_shift': user_shift,
        'user_next_access': user_next_access,
        'is_blocked': True
    }
    
    return render(request, 'production/shift_blocked.html', context)


# =====================================================
# TABLEAUX DE BORD
# =====================================================

@login_required
def admin_dashboard(request):
    if not (request.user.is_superuser or request.user.is_staff):
        messages.warning(request, 'Accès non autorisé')
        return redirect('operator_dashboard')

    
    try:
        today = timezone.now().date()
        today_records = ProductionRecord.objects.filter(date_heure__date=today)
        daily_kpis = calculate_kpis(today_records)
        
        # Statistiques principales
        stats = {
            'total_palettes': daily_kpis['total_palettes'],
            'total_palettes_conformes': daily_kpis['total_palettes_conformes'],
            'total_palettes_non_conformes': daily_kpis['total_palettes_non_conformes'],
            'total_dechets': daily_kpis['total_dechets'],
            'total_arrets': daily_kpis['total_arrets'],
            'nombre_enregistrements': daily_kpis['nombre_records'],
            'taux_productivite': daily_kpis['taux_productivite'],
            'taux_defauts': daily_kpis['taux_defauts'],
            'efficacite_globale': daily_kpis['efficacite_globale'],
            'palettes_par_heure': daily_kpis['palettes_par_heure'],
            'taux_conformite_global': daily_kpis['taux_conformite'],
            'temps_prevu_total': 480,
            'objectif_quotidien': 200,
        }
        # NOUVELLE SECTION : Récupérer les produits en cours de production
        now = timezone.now()  # obtient la date/heure actuelle
        active_threshold = now - timedelta(hours=2)
        
        # Récupérer les relevés récents avec production active
        recent_records = ProductionRecord.objects.filter(
            date_heure__gte=active_threshold,
            date_heure__date=today,
            palettes_produites__gt=0
        ).select_related('produit', 'operateur', 'ligne_production').order_by('-date_heure')
        
        # Grouper par combinaison produit/ligne pour éviter les doublons
        active_products = []
        seen_combinations = set()
        
        for record in recent_records:
            product_key = (
                record.produit.id if record.produit else 0, 
                record.ligne_production.id if record.ligne_production else 0
            )
            
            if product_key not in seen_combinations:
                # Calculer les totaux du jour pour cette combinaison
                day_totals = ProductionRecord.objects.filter(
                    produit=record.produit,
                    ligne_production=record.ligne_production,
                    date_heure__date=today
                ).aggregate(
                    total_palettes=Sum('palettes_produites'),
                    total_conformes=Sum('palettes_produites') - Sum('palettes_non_conformes'),
                    derniere_maj=Max('date_heure')
                )
                
                # Calculer le taux de conformité
                total_palettes = day_totals['total_palettes'] or 0
                total_conformes = day_totals['total_conformes'] or 0
                taux_conformite = (total_conformes / total_palettes * 100) if total_palettes > 0 else 100
                
                active_products.append({
                    'nom': record.produit.nom if record.produit else 'Produit inconnu',
                    'ligne_production': record.ligne_production,
                    'operateur': record.operateur,
                    'quantite_produite': total_palettes,
                    'conformes': total_conformes,
                    'taux_conformite': round(taux_conformite, 1),
                    'heure_debut': record.date_heure,
                    'derniere_maj': day_totals['derniere_maj'],
                    'duree_active': calculate_production_duration(record.date_heure, now)
                })
                seen_combinations.add(product_key)
        
        # Trier par quantité produite
        active_products.sort(key=lambda x: x['quantite_produite'], reverse=True)
        # Évolution par rapport à hier
        yesterday = today - timedelta(days=1)
        yesterday_records = ProductionRecord.objects.filter(date_heure__date=yesterday)
        yesterday_kpis = calculate_kpis(yesterday_records)
        
        if yesterday_kpis['total_palettes'] > 0:
            evolution_palettes = ((daily_kpis['total_palettes'] - yesterday_kpis['total_palettes']) / yesterday_kpis['total_palettes']) * 100
        else:
            evolution_palettes = 0
        
        stats['evolution_palettes'] = round(evolution_palettes, 2)
        
        # Données complémentaires
        recent_records = ProductionRecord.objects.select_related(
            'operateur', 'ligne_production', 'produit'
        ).order_by('-date_heure')[:10]
        
        alertes = Alerte.objects.filter(is_resolved=False).order_by('-date_creation')[:10]
        
        # Performance par ligne de production
        lignes_stats = []
        for ligne in LigneProduction.objects.filter(is_active=True):
            ligne_records = today_records.filter(ligne_production=ligne)
            ligne_kpis = calculate_kpis(ligne_records)
            
            lignes_stats.append({
                'ligne': ligne,
                'palettes': ligne_kpis['total_palettes'],
                'palettes_conformes': ligne_kpis['total_palettes_conformes'],
                'palettes_non_conformes': ligne_kpis['total_palettes_non_conformes'],
                'taux_conformite': ligne_kpis['taux_conformite'],
                'total_dechets': ligne_kpis['total_dechets'],
                'total_arrets': ligne_kpis['total_arrets'],
                'efficacite': ligne_kpis['efficacite_globale'],
                'taux_productivite': ligne_kpis['taux_productivite'],
                'palettes_par_heure': ligne_kpis['palettes_par_heure'],
                'derniere_activite': ligne_records.first().date_heure if ligne_records.exists() else None,
            })
        
        # Performance des équipes
        team_productivity = calculate_team_productivity()
        
        # KPIs hebdomadaires
        week_start = today - timedelta(days=7)
        weekly_records = ProductionRecord.objects.filter(date_heure__date__range=[week_start, today])
        weekly_kpis = calculate_kpis(weekly_records)
        
        # Nombre d'opérateurs actifs
        active_operators_count = ProductionRecord.objects.filter(
            date_heure__date=today
        ).values('operateur').distinct().count()
        
        # Informations sur l'équipe active
        active_shift, shift_name = determine_current_shift()
        current_hour = datetime.now().hour
        
        # Stats de l'équipe active
        active_records = ProductionRecord.objects.filter(
            date_heure__date=today,
            equipe__horaire=active_shift
        )
        
        context = {
            'stats': stats,
            'kpis': daily_kpis,
            'recent_records': recent_records,
            'alertes': alertes,
            'lignes_stats': lignes_stats,
            'team_productivity': team_productivity[:5],
            'current_time': timezone.now(),
            'weekly_kpis': weekly_kpis,
            'active_operators_count': active_operators_count,
            'active_shift': active_shift,
            'shift_name': shift_name,
            'current_hour': current_hour,
            'active_records_count': active_records.count(),
             'active_products': active_products[:6],  # Limiter à 6 pour l'affichage
            'active_products_count': len(active_products),
            'total_active_products': len(active_products),
        }
        
        return render(request, 'production/admin_dashboard.html', context)
        
    except Exception as e:
        messages.error(request, f'Erreur lors du chargement du tableau de bord: {str(e)}')
        return render(request, 'production/admin_dashboard.html', {
            'stats': {
                'total_palettes': 0, 'total_palettes_conformes': 0, 
                'total_palettes_non_conformes': 0, 'evolution_palettes': 0
            }, 
            'recent_records': [], 'alertes': [], 'lignes_stats': [],
            'team_productivity': [], 'kpis': {}, 'weekly_kpis': {},
            'current_time': timezone.now(), 'active_operators_count': 0
        })


@login_required
def operator_dashboard(request):
    """Dashboard opérateur avec contrôle d'accès par middleware"""
    try:
        user = request.user
        profile = user.userprofile if hasattr(user, 'userprofile') else None
        current_hour = datetime.now().hour
        
        # Stats de l'utilisateur pour aujourd'hui
        today = timezone.now().date()
        user_records = ProductionRecord.objects.filter(
            operateur=user,
            date_heure__date=today
        ).order_by('-date_heure')[:10]
        
        # Calcul des KPIs utilisateur
        today_records = ProductionRecord.objects.filter(operateur=user, date_heure__date=today)
        user_kpis = calculate_kpis(today_records)
        
        user_stats = {
            'palettes_today': user_kpis['total_palettes'],
            'dechets_today': user_kpis['total_dechets'],
            'arrets_today': user_kpis['total_arrets'],
            'taux_productivite': user_kpis['taux_productivite'],
            'efficacite_globale': user_kpis['efficacite_globale'],
            'palettes_par_heure': user_kpis['palettes_par_heure'],
            'taux_conformite': user_kpis['taux_conformite'],
            'palettes_conformes': user_kpis['total_palettes_conformes'],
            'palettes_non_conformes': user_kpis['total_palettes_non_conformes'],
        }
        
        # Informations sur l'équipe
        current_shift, shift_name = determine_current_shift()
        
        context = {
            'recent_records': user_records,
            'user_stats': user_stats,
            'current_shift': profile.horaire_equipe if profile else current_shift,
            'current_hour': current_hour,
            'shift_name': profile.get_horaire_equipe_display() if profile else shift_name,
            'current_time': timezone.now(),
        }
        
        return render(request, 'production/operator_dashboard.html', context)
        
    except Exception as e:
        messages.error(request, f'Erreur lors du chargement de l\'interface: {str(e)}')
        return render(request, 'production/operator_dashboard.html', {
            'recent_records': [], 'user_stats': {}, 'current_time': timezone.now()
        })


# =====================================================
# GESTION DES ENREGISTREMENTS DE PRODUCTION
# =====================================================

@login_required
@csrf_protect
def add_production_record(request):
    """Créer un enregistrement de production avec validation d'équipe"""
    if request.method == 'POST':
        form = ProductionForm(request.POST)
        if form.is_valid():
            try:
                record = form.save(commit=False)
                # ✅ CORRECTION - assigner l'opérateur AVANT toute validation
                record.operateur = request.user
                
                # Déterminer l'horaire actuel
                current_shift, shift_name = determine_current_shift()
                
                # Créer ou récupérer l'équipe
                equipe, created = Equipe.objects.get_or_create(
                    operateur=request.user,
                    horaire=current_shift,
                    defaults={
                        'nom': f"Équipe {current_shift} - {request.user.username}",
                        'is_active': True
                    }
                )
                record.equipe = equipe
                
                # Maintenant valider avec l'opérateur assigné
                record.full_clean()  # Cette ligne déclenche clean()
                record.save()
    
                
                # Initialiser les champs de traitement NC
                if hasattr(record, 'statut_palette_nc'):
                    if record.palettes_non_conformes > 0:
                        record.statut_palette_nc = 'en_attente'
                    else:
                        record.statut_palette_nc = 'conforme'
                
                record.save()
                
                # Créer des alertes automatiques
                if hasattr(record, 'duree_arret_minutes') and record.duree_arret_minutes and record.duree_arret_minutes > 5:
                    Alerte.objects.create(
                        production_record=record,
                        type_alerte='arret',
                        message=f"Arrêt de {record.duree_arret_minutes} minutes sur {record.ligne_production}. Cause: {getattr(record, 'cause_arret', '') or 'Non spécifiée'}"
                    )
                
                if record.dechets_boites and record.dechets_boites > 10:
                    Alerte.objects.create(
                        production_record=record,
                        type_alerte='dechets',
                        message=f"Déchets élevés détectés : {record.dechets_boites} boites sur {record.ligne_production}"
                    )
                    messages.warning(
                        request,
                        f"Attention : Niveau de déchets élevé ({record.dechets_boites} boites)"
                    )
                
                messages.success(request, 'Enregistrement créé avec succès!')
                return redirect('operator_dashboard')
                
            except Exception as e:
                messages.error(request, f'Erreur lors de l\'enregistrement: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = ProductionForm()
    
    return render(request, 'production/operator_form.html', {'form': form})


@login_required
def edit_production_record(request, record_id):
    """Modifier un enregistrement de production (dans les 2 heures)"""
    record = get_object_or_404(ProductionRecord, id=record_id, operateur=request.user)
    
    # Vérifier la limite de temps pour la modification
    time_limit = timezone.now() - timedelta(hours=2)
    if record.date_heure < time_limit:
        messages.error(request, "Ce relevé ne peut plus être modifié (plus de 2 heures).")
        return redirect('production_history')
    
    if request.method == 'POST':
        form = ProductionForm(request.POST, instance=record)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Relevé modifié avec succès!')
                return redirect('production_history')
            except Exception as e:
                messages.error(request, f'Erreur lors de la modification: {str(e)}')
    else:
        form = ProductionForm(instance=record)
    
    return render(request, 'production/edit_production_record.html', {
        'form': form, 
        'record': record
    })


# =====================================================
# HISTORIQUE ET RAPPORTS
# =====================================================

@login_required
def production_history(request):
    """Vue pour afficher l'historique de production avec filtres et traitement NC"""
    
    # Récupérer tous les enregistrements
    records = ProductionRecord.objects.select_related(
        'operateur', 'ligne_production', 'produit', 'equipe'
    ).order_by('-date_heure')
    
    # Appliquer les filtres
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    operateur_id = request.GET.get('operateur')
    ligne_id = request.GET.get('ligne')
    
    if date_debut:
        try:
            date_debut_parsed = timezone.strptime(date_debut, '%Y-%m-%d').date()
            records = records.filter(date_heure__date__gte=date_debut_parsed)
        except ValueError:
            messages.warning(request, "Format de date de début incorrect")
    
    if date_fin:
        try:
            date_fin_parsed = timezone.strptime(date_fin, '%Y-%m-%d').date()
            records = records.filter(date_heure__date__lte=date_fin_parsed)
        except ValueError:
            messages.warning(request, "Format de date de fin incorrect")
    
    if operateur_id:
        records = records.filter(operateur_id=operateur_id)
    
    if ligne_id:
        records = records.filter(ligne_production_id=ligne_id)
    
    # Gestion de l'export
    export_type = request.GET.get('export')
    if export_type in ['csv', 'excel']:
        return export_production_data(records, export_type)
    
    # Pagination
    paginator = Paginator(records, 20)
    page_number = request.GET.get('page')
    page_records = paginator.get_page(page_number)
    
    # Calculer et ajouter des informations supplémentaires à chaque enregistrement
    for record in page_records:
        # Calcul du taux de conformité
        if record.palettes_produites > 0:
            record.taux_conformite = ((record.palettes_produites - record.palettes_non_conformes) / record.palettes_produites) * 100
        else:
            record.taux_conformite = 0
        
        # S'assurer que le statut NC existe et est correct
        if not hasattr(record, 'statut_palette_nc') or not record.statut_palette_nc:
            if record.palettes_non_conformes > 0:
                record.statut_palette_nc = 'en_attente'
                record.save()
            else:
                record.statut_palette_nc = 'conforme'
                record.save()
        
        # Temps d'arrêt
        record.temps_arret_value = getattr(record, 'duree_arret_minutes', 0) or 0
    
    # Statistiques du jour
    today = timezone.now().date()
    today_aggregation = ProductionRecord.objects.filter(
        date_heure__date=today
    ).aggregate(
        total_palettes=Sum('palettes_produites'),
        total_palettes_non_conformes=Sum('palettes_non_conformes'),
        total_dechets=Sum('dechets_boites')
    )
    
    # Calcul des statistiques sécurisé
    total_palettes = today_aggregation['total_palettes'] or 0
    total_palettes_nc = today_aggregation['total_palettes_non_conformes'] or 0
    total_dechets = today_aggregation['total_dechets'] or 0
    
    today_stats = {
        'total_palettes': total_palettes,
        'total_palettes_conformes': total_palettes - total_palettes_nc,
        'total_palettes_non_conformes': total_palettes_nc,
        'total_dechets': float(total_dechets)
    }
    
    # Données pour les filtres
    operateurs = User.objects.filter(is_active=True).order_by('username')
    lignes = LigneProduction.objects.filter(is_active=True).order_by('numero')
    
    context = {
        'records': page_records,
        'today_stats': today_stats,
        'today': today,
        'operateurs': operateurs,
        'lignes': lignes,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'operateur_id': operateur_id,
        'ligne_id': ligne_id,
    }
    
    return render(request, 'production/production_history.html', context)


# =====================================================
# GESTION DES PALETTES NON CONFORMES
# =====================================================

@login_required
@user_passes_test(lambda u: u.is_staff)
def palette_nc_details(request, record_id):
    """API pour récupérer les détails d'un enregistrement avec palettes NC"""
    
    try:
        record = get_object_or_404(
            ProductionRecord.objects.select_related('operateur', 'ligne_production', 'produit'),
            id=record_id
        )
        
        data = {
            'id': record.id,
            'produit': record.produit.nom if record.produit else 'N/A',
            'operateur': record.operateur.username if record.operateur else 'N/A',
            'ligne': f"Ligne {record.ligne_production.numero}" if record.ligne_production else 'N/A',
            'date_heure': record.date_heure.strftime('%d/%m/%Y à %H:%M'),
            'palettes_nc': record.palettes_non_conformes,
            'palettes_produites': record.palettes_produites,
            'cause_non_conformite': getattr(record, 'cause_non_conformite', '') or 'Non renseignée',
            'statut': getattr(record, 'statut_palette_nc', 'en_attente'),
            'dechets_actuels': record.dechets_boites or 0
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@user_passes_test(lambda u: u.is_staff)
@csrf_protect
def traiter_palette_nc(request, record_id):
    """Traiter les palettes non conformes avec gestion des déchets améliorée"""
    
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée')
        return redirect('production_history')
    
    try:
        record = get_object_or_404(ProductionRecord, id=record_id)
        
        # Vérifier qu'il y a bien des palettes NC à traiter
        if record.palettes_non_conformes <= 0:
            messages.error(request, "Aucune palette non conforme à traiter.")
            return redirect('production_history')
        
        action = request.POST.get('action')
        dechets_initiaux = record.dechets_boites or 0  # Sauvegarder les déchets initiaux
        
        if action == 'toute_conforme':
            # Toutes les palettes NC deviennent conformes
            palettes_recuperees = record.palettes_non_conformes
            
            # Réajuster les compteurs
            record.palettes_non_conformes = 0
            
            # Mettre à jour les champs de traçabilité
            if hasattr(record, 'statut_palette_nc'):
                record.statut_palette_nc = 'traitee'
            if hasattr(record, 'date_controle_nc'):
                record.date_controle_nc = timezone.now()
            if hasattr(record, 'commentaire_controle_nc'):
                record.commentaire_controle_nc = f"Contrôle qualité: {palettes_recuperees} palette(s) validée(s) comme conformes - Traité par {request.user.username}"
            if hasattr(record, 'fardeaux_recuperes'):
                record.fardeaux_recuperes = palettes_recuperees * 72  # 1 palette = 72 fardeaux
            
            messages.success(request, f"✅ {palettes_recuperees} palette(s) validée(s) comme conformes!")
            
        elif action == 'traitement_detaille':
            # Traitement détaillé avec gestion précise des déchets
            palettes_conformes = int(request.POST.get('palettes_conformes', 0))
            fardeaux_partiels = int(request.POST.get('fardeaux_partiels', 0))
            commentaire = request.POST.get('commentaire_controle_detaille', '')
            
            # Validation des données
            if palettes_conformes < 0 or fardeaux_partiels < 0:
                raise ValueError("Les valeurs ne peuvent pas être négatives")
            
            if palettes_conformes > record.palettes_non_conformes:
                raise ValueError("Le nombre de palettes récupérées ne peut pas dépasser les palettes non conformes")
            
            # Calculs selon la logique : 1 palette = 72 fardeaux, 1 fardeau = 24 boîtes
            total_fardeaux_dans_nc = record.palettes_non_conformes * 72
            fardeaux_des_palettes_conformes = palettes_conformes * 72
            
            # Vérifier que les fardeaux partiels ne dépassent pas ce qui reste
            fardeaux_restants_disponibles = total_fardeaux_dans_nc - fardeaux_des_palettes_conformes
            if fardeaux_partiels > fardeaux_restants_disponibles:
                fardeaux_partiels = fardeaux_restants_disponibles
                messages.warning(request, f"Fardeaux ajustés à {fardeaux_partiels} (maximum disponible)")
            
            # Calculer les fardeaux non récupérés → déchets
            fardeaux_non_recuperes = total_fardeaux_dans_nc - fardeaux_des_palettes_conformes - fardeaux_partiels
            boites_dechets_supplementaires = fardeaux_non_recuperes * 24  # 1 fardeau = 24 boîtes
            
            # Mettre à jour les déchets
            record.dechets_boites = dechets_initiaux + boites_dechets_supplementaires
            
            # Mettre à jour les compteurs (toutes les palettes NC ont été traitées)
            record.palettes_non_conformes = 0
            
            # Mettre à jour les champs de traçabilité
            if hasattr(record, 'statut_palette_nc'):
                record.statut_palette_nc = 'traitee'
            if hasattr(record, 'date_controle_nc'):
                record.date_controle_nc = timezone.now()
            if hasattr(record, 'commentaire_controle_nc'):
                record.commentaire_controle_nc = f"Traitement détaillé: {palettes_conformes} palette(s) entières + {fardeaux_partiels} fardeaux partiels récupérés. {fardeaux_non_recuperes} fardeaux → {boites_dechets_supplementaires} boîtes de déchets. {commentaire} - Traité par {request.user.username}"
            if hasattr(record, 'fardeaux_recuperes'):
                record.fardeaux_recuperes = fardeaux_des_palettes_conformes + fardeaux_partiels
            
            messages.success(request, f"✅ Traitement terminé: {palettes_conformes} palette(s) + {fardeaux_partiels} fardeaux récupérés, {boites_dechets_supplementaires} boîtes ajoutées aux déchets")
            
        elif action == 'partiel':
            # Traitement rapide partiel avec conversion correcte 1 fardeau = 24 boîtes
            fardeaux_conformes = int(request.POST.get('fardeaux_conformes', 0))
            commentaire = request.POST.get('commentaire_controle', '')
            
            if fardeaux_conformes <= 0:
                raise ValueError("Le nombre de fardeaux doit être positif")
            
            # Calculs selon la logique correcte
            total_fardeaux_dans_nc = record.palettes_non_conformes * 72  # 1 palette = 72 fardeaux
            
            if fardeaux_conformes > total_fardeaux_dans_nc:
                raise ValueError(f"Le nombre de fardeaux ne peut pas dépasser {total_fardeaux_dans_nc}")
            
            # Calculer les fardeaux non récupérés → déchets (1 fardeau = 24 boîtes)
            fardeaux_non_recuperes = total_fardeaux_dans_nc - fardeaux_conformes
            boites_dechets_supplementaires = fardeaux_non_recuperes * 24
            
            # Mettre à jour les déchets
            record.dechets_boites = dechets_initiaux + boites_dechets_supplementaires
            
            # Toutes les palettes NC ont été traitées (soit récupérées, soit devenues déchets)
            record.palettes_non_conformes = 0
            
            # Mettre à jour les champs de traçabilité
            if hasattr(record, 'statut_palette_nc'):
                record.statut_palette_nc = 'traitee'
                    
            if hasattr(record, 'date_controle_nc'):
                record.date_controle_nc = timezone.now()
                
            commentaire_final = f"Récupération partielle: {fardeaux_conformes} fardeau(s) récupéré(s), {fardeaux_non_recuperes} fardeaux non récupérés = {boites_dechets_supplementaires} boîtes de déchets"
            if commentaire:
                commentaire_final += f" - {commentaire}"
            commentaire_final += f" - Traité par {request.user.username}"
            
            if hasattr(record, 'commentaire_controle_nc'):
                record.commentaire_controle_nc = commentaire_final
                
            if hasattr(record, 'fardeaux_recuperes'):
                record.fardeaux_recuperes = (record.fardeaux_recuperes or 0) + fardeaux_conformes
            
            messages.success(request, f"✅ {fardeaux_conformes} fardeau(s) récupéré(s), {boites_dechets_supplementaires} boîtes ajoutées aux déchets")
        
        elif action == 'non_conforme':
            # Marquer toutes comme définitivement non conformes (déchets)
            palettes_rejetees = record.palettes_non_conformes
            
            # Ajouter les palettes rejetées aux déchets (1 palette = 72 boîtes)
            boites_dechets_supplementaires = palettes_rejetees * 72
            record.dechets_boites = dechets_initiaux + boites_dechets_supplementaires
            
            # Réinitialiser les palettes NC (elles sont maintenant comptées comme déchets)
            record.palettes_non_conformes = 0
            
            # Mettre à jour les champs de traçabilité
            if hasattr(record, 'statut_palette_nc'):
                record.statut_palette_nc = 'rejetee'
            if hasattr(record, 'date_controle_nc'):
                record.date_controle_nc = timezone.now()
            if hasattr(record, 'commentaire_controle_nc'):
                record.commentaire_controle_nc = f"Contrôle qualité: {palettes_rejetees} palette(s) confirmée(s) non conformes → {boites_dechets_supplementaires} boîtes en déchets - Traité par {request.user.username}"
            
            messages.warning(request, f"⚠️ {palettes_rejetees} palette(s) rejetée(s) = {boites_dechets_supplementaires} boîtes ajoutées aux déchets.")
        
        else:
            messages.error(request, "Action non reconnue.")
            return redirect('production_history')
        
        # Sauvegarder les modifications
        record.save()
        
        # Log pour traçabilité
        print(f"[PALETTE_NC] Record ID: {record_id}, Action: {action}, User: {request.user.username}, "
              f"Déchets avant: {dechets_initiaux}, Déchets après: {record.dechets_boites}, Timestamp: {timezone.now()}")
        
        return redirect('production_history')
        
    except Exception as e:
        messages.error(request, f"Erreur lors du traitement: {str(e)}")
        return redirect('production_history')


# =====================================================
# FONCTIONS D'EXPORT
# =====================================================

def export_production_data(queryset, export_type):
    """Exporter les données de production en CSV ou Excel avec colonne déchets"""
    
    if export_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="production_history_{timezone.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Date/Heure', 'Opérateur', 'Ligne', 'Produit', 
            'Palettes Produites', 'Palettes NC', 'Déchets (boîtes)', 'Taux Conformité (%)',
            'Cause NC', 'Fardeaux Récupérés', 'Temps Arrêt (min)', 'Cause Arrêt'
        ])
        
        for record in queryset:
            taux_conformite = 0
            if record.palettes_produites > 0:
                taux_conformite = ((record.palettes_produites - record.palettes_non_conformes) / record.palettes_produites) * 100
            
            writer.writerow([
                record.date_heure.strftime('%d/%m/%Y %H:%M'),
                record.operateur.username if record.operateur else 'N/A',
                f"L{record.ligne_production.numero}" if record.ligne_production else 'N/A',
                record.produit.nom if record.produit else 'N/A',
                record.palettes_produites,
                record.palettes_non_conformes,
                record.dechets_boites or 0,  # Nouvelle colonne déchets
                f"{taux_conformite:.1f}",
                getattr(record, 'cause_non_conformite', '') or '',
                getattr(record, 'fardeaux_recuperes', 0) or 0,
                getattr(record, 'duree_arret_minutes', 0) or 0,
                getattr(record, 'cause_arret', '') or ''
            ])
        
        return response
    
    elif export_type == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="production_history_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Historique Production"
        
        # En-têtes avec style
        headers = [
            'Date/Heure', 'Opérateur', 'Ligne', 'Produit', 
            'Palettes Produites', 'Palettes NC', 'Déchets (boîtes)', 'Taux Conformité (%)',
            'Cause NC', 'Fardeaux Récupérés', 'Temps Arrêt (min)', 'Cause Arrêt'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Données
        for row_num, record in enumerate(queryset, 2):
            taux_conformite = 0
            if record.palettes_produites > 0:
                taux_conformite = ((record.palettes_produites - record.palettes_non_conformes) / record.palettes_produites) * 100
            
            worksheet.cell(row=row_num, column=1).value = record.date_heure.strftime('%d/%m/%Y %H:%M')
            worksheet.cell(row=row_num, column=2).value = record.operateur.username if record.operateur else 'N/A'
            worksheet.cell(row=row_num, column=3).value = f"L{record.ligne_production.numero}" if record.ligne_production else 'N/A'
            worksheet.cell(row=row_num, column=4).value = record.produit.nom if record.produit else 'N/A'
            worksheet.cell(row=row_num, column=5).value = record.palettes_produites
            worksheet.cell(row=row_num, column=6).value = record.palettes_non_conformes
            worksheet.cell(row=row_num, column=7).value = record.dechets_boites or 0  # Nouvelle colonne déchets
            worksheet.cell(row=row_num, column=8).value = f"{taux_conformite:.1f}"
            worksheet.cell(row=row_num, column=9).value = getattr(record, 'cause_non_conformite', '') or ''
            worksheet.cell(row=row_num, column=10).value = getattr(record, 'fardeaux_recuperes', 0) or 0
            worksheet.cell(row=row_num, column=11).value = getattr(record, 'duree_arret_minutes', 0) or 0
            worksheet.cell(row=row_num, column=12).value = getattr(record, 'cause_arret', '') or ''
        
        # Ajuster la largeur des colonnes
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
        
        workbook.save(response)
        return response


# =====================================================
# APIs ET FONCTIONS UTILITAIRES
# =====================================================

@login_required
def api_production_data(request):
    """API pour récupérer les données de production en temps réel avec produits actifs"""
    if not (request.user.is_superuser or request.user.is_staff):
        return JsonResponse({'error': 'Non autorisé'}, status=403)
    
    try:
        now = timezone.now()
        today = now.date()
        
        # Définir la période "active" (relevés des 2 dernières heures)
        active_threshold = now - timedelta(hours=2)
        
        # Récupérer les relevés récents avec production active
        recent_records = ProductionRecord.objects.filter(
            date_heure__gte=active_threshold,
            date_heure__date=today,
            palettes_produites__gt=0  # Seulement les productions avec quantité > 0
        ).select_related('produit', 'operateur', 'ligne_production').order_by('-date_heure')
        
        # Grouper par combinaison produit/ligne pour éviter les doublons
        active_products = []
        seen_combinations = set()
        
        for record in recent_records:
            # Créer une clé unique basée sur produit + ligne
            product_key = (
                record.produit.id if record.produit else 0, 
                record.ligne_production.id if record.ligne_production else 0
            )
            
            if product_key not in seen_combinations:
                # Calculer les totaux du jour pour cette combinaison
                day_totals = ProductionRecord.objects.filter(
                    produit=record.produit,
                    ligne_production=record.ligne_production,
                    date_heure__date=today
                ).aggregate(
                    total_palettes=Sum('palettes_produites'),
                    total_conformes=Sum('palettes_produites') - Sum('palettes_non_conformes'),
                    derniere_maj=Max('date_heure')
                )
                
                # Trouver le dernier opérateur actif sur cette combinaison
                last_operator = recent_records.filter(
                    produit=record.produit,
                    ligne_production=record.ligne_production
                ).first().operateur
                
                active_products.append({
                    'id': f"{record.produit.id if record.produit else 0}_{record.ligne_production.id if record.ligne_production else 0}",
                    'nom': record.produit.nom if record.produit else 'Produit inconnu',
                    'ligne': f"Ligne {record.ligne_production.numero}" if record.ligne_production else 'Ligne inconnue',
                    'operateur': {
                        'nom': last_operator.username if last_operator else 'N/A',
                        'full_name': f"{last_operator.first_name} {last_operator.last_name}".strip() if last_operator and last_operator.first_name else last_operator.username if last_operator else 'N/A'
                    },
                    'quantite_produite': day_totals['total_palettes'] or 0,
                    'quantite_conforme': day_totals['total_conformes'] or 0,
                    'heure_debut': record.date_heure.strftime('%H:%M'),
                    'derniere_maj': day_totals['derniere_maj'].strftime('%H:%M:%S') if day_totals['derniere_maj'] else 'N/A',
                    'statut': 'production',
                    'duree_production': self.calculate_production_duration(record.date_heure, now)
                })
                seen_combinations.add(product_key)
        
        # Trier par quantité produite (descendant)
        active_products.sort(key=lambda x: x['quantite_produite'], reverse=True)
        
        # Compter les opérateurs actifs
        active_operators_count = recent_records.values('operateur').distinct().count()
        
        # Statistiques KPI générales
        last_24h = now - timedelta(hours=24)
        records = ProductionRecord.objects.filter(date_heure__gte=last_24h)
        kpis = calculate_kpis(records)
        
        data = {
            'active_products_count': len(active_products),
            'active_products': active_products[:10],  # Limiter à 10 pour l'affichage
            'active_operators_count': active_operators_count,
            'current_time': now.strftime('%H:%M:%S'),
            'last_update': now.isoformat(),
            'products_html': generate_products_html(active_products[:6]),  # HTML pour 6 produits max
            
            # KPIs généraux (pour compatibilité)
            'total_palettes': kpis['total_palettes'],
            'total_dechets': kpis['total_dechets'],
            'alertes_count': Alerte.objects.filter(is_resolved=False).count(),
            'taux_productivite': kpis['taux_productivite'],
            'efficacite_globale': kpis['efficacite_globale'],
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'active_products_count': 0,
            'active_products': [],
            'current_time': timezone.now().strftime('%H:%M:%S'),
            'last_update': timezone.now().isoformat(),
            'products_html': '<div class="text-center text-muted py-3"><i class="fas fa-exclamation-triangle"></i><p>Erreur de chargement</p></div>'
        }, status=500)
@login_required
def api_production_data(request):
    """API pour récupérer les données de production en temps réel avec produits actifs"""
    if not (request.user.is_superuser or request.user.is_staff):
        return JsonResponse({'error': 'Non autorisé'}, status=403)
    
    try:
        now = timezone.now()
        today = now.date()
        
        # Définir la période "active" (relevés des 2 dernières heures)
        active_threshold = now - timedelta(hours=2)
        
        # Récupérer les relevés récents avec production active
        recent_records = ProductionRecord.objects.filter(
            date_heure__gte=active_threshold,
            date_heure__date=today,
            palettes_produites__gt=0  # Seulement les productions avec quantité > 0
        ).select_related('produit', 'operateur', 'ligne_production').order_by('-date_heure')
        
        # Grouper par combinaison produit/ligne pour éviter les doublons
        active_products = []
        seen_combinations = set()
        
        for record in recent_records:
            # Créer une clé unique basée sur produit + ligne
            product_key = (
                record.produit.id if record.produit else 0, 
                record.ligne_production.id if record.ligne_production else 0
            )
            
            if product_key not in seen_combinations:
                # Calculer les totaux du jour pour cette combinaison
                day_totals = ProductionRecord.objects.filter(
                    produit=record.produit,
                    ligne_production=record.ligne_production,
                    date_heure__date=today
                ).aggregate(
                    total_palettes=Sum('palettes_produites'),
                    total_conformes=Sum('palettes_produites') - Sum('palettes_non_conformes'),
                    derniere_maj=Max('date_heure')
                )
                
                # Trouver le dernier opérateur actif sur cette combinaison
                last_operator = recent_records.filter(
                    produit=record.produit,
                    ligne_production=record.ligne_production
                ).first().operateur
                
                active_products.append({
                    'id': f"{record.produit.id if record.produit else 0}_{record.ligne_production.id if record.ligne_production else 0}",
                    'nom': record.produit.nom if record.produit else 'Produit inconnu',
                    'ligne': f"Ligne {record.ligne_production.numero}" if record.ligne_production else 'Ligne inconnue',
                    'operateur': {
                        'nom': last_operator.username if last_operator else 'N/A',
                        'full_name': f"{last_operator.first_name} {last_operator.last_name}".strip() if last_operator and last_operator.first_name else last_operator.username if last_operator else 'N/A'
                    },
                    'quantite_produite': day_totals['total_palettes'] or 0,
                    'quantite_conforme': day_totals['total_conformes'] or 0,
                    'heure_debut': record.date_heure.strftime('%H:%M'),
                    'derniere_maj': day_totals['derniere_maj'].strftime('%H:%M:%S') if day_totals['derniere_maj'] else 'N/A',
                    'statut': 'production',
                    'duree_production': calculate_production_duration(record.date_heure, now)
                })
                seen_combinations.add(product_key)
        
        # Trier par quantité produite (descendant)
        active_products.sort(key=lambda x: x['quantite_produite'], reverse=True)
        
        # Compter les opérateurs actifs
        active_operators_count = recent_records.values('operateur').distinct().count()
        
        # Statistiques KPI générales
        last_24h = now - timedelta(hours=24)
        records = ProductionRecord.objects.filter(date_heure__gte=last_24h)
        kpis = calculate_kpis(records)
        
        data = {
            'active_products_count': len(active_products),
            'active_products': active_products[:10],  # Limiter à 10 pour l'affichage
            'active_operators_count': active_operators_count,
            'current_time': now.strftime('%H:%M:%S'),
            'last_update': now.isoformat(),
            'products_html': generate_products_html(active_products[:6]),  # HTML pour 6 produits max
            
            # KPIs généraux (pour compatibilité)
            'total_palettes': kpis['total_palettes'],
            'total_dechets': kpis['total_dechets'],
            'alertes_count': Alerte.objects.filter(is_resolved=False).count(),
            'taux_productivite': kpis['taux_productivite'],
            'efficacite_globale': kpis['efficacite_globale'],
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'active_products_count': 0,
            'active_products': [],
            'current_time': timezone.now().strftime('%H:%M:%S'),
            'last_update': timezone.now().isoformat(),
            'products_html': '<div class="text-center text-muted py-3"><i class="fas fa-exclamation-triangle"></i><p>Erreur de chargement</p></div>'
        }, status=500)

def calculate_production_duration(start_time, current_time):
    """Calcule la durée de production en format lisible"""
    try:
        delta = current_time - start_time
        hours = int(delta.total_seconds() // 3600)
        minutes = int((delta.total_seconds() % 3600) // 60)
        
        if hours > 0:
            return f"{hours}h {minutes:02d}min"
        else:
            return f"{minutes}min"
    except:
        return "N/A"



@login_required
def resolve_alert(request, alert_id):
    """Résoudre une alerte"""
    if not (request.user.is_superuser or request.user.is_staff):
        messages.warning(request, 'Accès non autorisé')
        return redirect('operator_dashboard')
    
    try:
        alert = get_object_or_404(Alerte, id=alert_id)
        alert.is_resolved = True
        alert.date_resolution = timezone.now()
        alert.save()
        messages.success(request, 'Alerte marquée comme résolue')
        
    except Exception as e:
        messages.error(request, f'Erreur lors de la résolution de l\'alerte: {str(e)}')
    
    return redirect('admin_dashboard')


@login_required
def api_system_health(request):
    """API pour vérifier la santé du système"""
    if not (request.user.is_superuser or request.user.is_staff):
        return JsonResponse({'error': 'Non autorisé'}, status=403)
    
    try:
        today = timezone.now().date()
        
        records_count = ProductionRecord.objects.filter(date_heure__date=today).count()
        unresolved_alerts = Alerte.objects.filter(is_resolved=False).count()
        active_users = ProductionRecord.objects.filter(
            date_heure__date=today
        ).values('operateur').distinct().count()
        active_lines = ProductionRecord.objects.filter(
            date_heure__date=today
        ).values('ligne_production').distinct().count()
        
        # Calcul du score de santé
        health_score = 100
        
        if unresolved_alerts > 10:
            health_score -= 20
        elif unresolved_alerts > 5:
            health_score -= 10
        
        if records_count == 0:
            health_score -= 30
        elif records_count < 10:
            health_score -= 15
        
        # Statut basé sur le score
        if health_score >= 90:
            status = 'excellent'
        elif health_score >= 70:
            status = 'good'
        elif health_score >= 50:
            status = 'warning'
        else:
            status = 'critical'
        
        data = {
            'status': status,
            'health_score': health_score,
            'metrics': {
                'records_today': records_count,
                'unresolved_alerts': unresolved_alerts,
                'active_users': active_users,
                'active_lines': active_lines
            },
            'timestamp': timezone.now().isoformat(),
            'database_status': 'connected',
            'last_activity': ProductionRecord.objects.all().order_by('-date_heure').first().date_heure.isoformat() if ProductionRecord.objects.exists() else None
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'error': str(e),
            'timestamp': timezone.now().isoformat()
        }, status=500)
    # =====================================================
# NOUVELLES FONCTIONS POUR L'INDICATEUR DE PRODUCTION
# =====================================================

def get_active_production_status():
    """Fonction utilitaire pour récupérer les produits en cours de production"""
    now = timezone.now()
    today = now.date()
    
    # Définir la période "active" (relevés des 2 dernières heures)
    active_threshold = now - timedelta(hours=2)
    
    # Récupérer les relevés récents avec production active
    active_records = ProductionRecord.objects.filter(
        date_heure__gte=active_threshold,
        date_heure__date=today,
        palettes_produites__gt=0  # Produits avec une quantité > 0
    ).select_related('produit', 'operateur', 'ligne_production').order_by('-date_heure')
    
    # Grouper par produit/ligne pour éviter les doublons
    active_products = []
    seen_combinations = set()
    
    for record in active_records:
        # Créer une clé unique basée sur produit + ligne
        product_key = (
            record.produit.id if record.produit else 0, 
            record.ligne_production.id if record.ligne_production else 0
        )
        
        if product_key not in seen_combinations:
            # Calculer la quantité totale produite aujourd'hui pour cette combinaison
            total_produit = ProductionRecord.objects.filter(
                produit=record.produit,
                ligne_production=record.ligne_production,
                date_heure__date=today
            ).aggregate(
                total_palettes=Sum('palettes_produites'),
                total_conformes=Sum('palettes_produites') - Sum('palettes_non_conformes'),
                derniere_maj=Max('date_heure')
            )
            
            active_products.append({
                'id': f"{record.produit.id if record.produit else 0}_{record.ligne_production.id if record.ligne_production else 0}",
                'produit_nom': record.produit.nom if record.produit else 'Produit inconnu',
                'ligne_nom': f"Ligne {record.ligne_production.numero}" if record.ligne_production else 'Ligne inconnue',
                'operateur': record.operateur,
                'quantite_produite': total_produit['total_palettes'] or 0,
                'quantite_conforme': total_produit['total_conformes'] or 0,
                'heure_debut': record.date_heure,
                'derniere_maj': total_produit['derniere_maj'],
                'statut': 'production'
            })
            seen_combinations.add(product_key)
    
    # Compter les opérateurs actifs
    active_operators = ProductionRecord.objects.filter(
        date_heure__gte=active_threshold,
        date_heure__date=today
    ).values('operateur').distinct().count()
    
    return {
        'active_products': active_products,
        'active_products_count': len(active_products),
        'active_operators_count': active_operators,
        'current_time': now,
        'last_update': now
    }


def generate_products_html(active_products):
    """Générer le HTML pour les produits actifs avec design amélioré"""
    if not active_products:
        return '''
        <div class="col-12">
            <div class="text-center text-muted py-4">
                <i class="fas fa-pause-circle fa-3x mb-3 text-muted"></i>
                <h6 class="text-muted">Aucune production active</h6>
                <small>Les relevés des 2 dernières heures s'afficheront ici</small>
            </div>
        </div>
        '''
    
    html = ''
    for product in active_products:
        # Calculer le pourcentage de conformité
        conformity_percent = 0
        if product['quantite_produite'] > 0:
            conformity_percent = (product['quantite_conforme'] / product['quantite_produite']) * 100
        
        # Déterminer la couleur du badge selon la performance
        badge_color = 'success'
        if conformity_percent < 90:
            badge_color = 'warning'
        if conformity_percent < 80:
            badge_color = 'danger'
        
        html += f'''
        <div class="col-md-6 col-lg-4 mb-3">
            <div class="card border-0 shadow-sm bg-light hover-card">
                <div class="card-body p-3">
                    <div class="d-flex justify-content-between align-items-start">
                        <div class="flex-grow-1">
                            <h6 class="text-primary mb-1 fw-bold">
                                <i class="fas fa-box-open text-primary me-1"></i>
                                {product['nom'][:20]}{'...' if len(product['nom']) > 20 else ''}
                            </h6>
                            <div class="text-muted small mb-2">
                                <div><i class="fas fa-industry me-1"></i> {product['ligne']}</div>
                                <div><i class="fas fa-user me-1"></i> {product['operateur']['full_name']}</div>
                                <div><i class="fas fa-clock me-1"></i> Depuis {product['heure_debut']}</div>
                            </div>
                        </div>
                        <div class="text-center ms-2">
                            <span class="badge bg-{badge_color} fs-6 px-2 py-1">
                                {product['quantite_produite']}
                            </span>
                            <div class="small text-muted mt-1">palettes</div>
                            <div class="progress mt-2" style="height: 4px; width: 60px;">
                                <div class="progress-bar bg-{badge_color} progress-bar-striped progress-bar-animated" 
                                     style="width: {min(100, conformity_percent)}%"></div>
                            </div>
                            <small class="text-muted">{conformity_percent:.0f}% OK</small>
                        </div>
                    </div>
                    
                    <div class="row mt-2 pt-2 border-top">
                        <div class="col-6 text-center">
                            <small class="text-success fw-bold">{product['quantite_conforme']}</small>
                            <div class="text-muted" style="font-size: 0.75rem;">Conformes</div>
                        </div>
                        <div class="col-6 text-center">
                            <small class="text-muted">{product['derniere_maj']}</small>
                            <div class="text-muted" style="font-size: 0.75rem;">Dernière MAJ</div>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        '''
    
    return html
